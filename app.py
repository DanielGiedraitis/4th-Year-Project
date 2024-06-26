from flask import Flask, render_template, request, redirect, url_for, session, jsonify
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime
import nltk
from nltk.tokenize import word_tokenize
import openpyxl
from openai import OpenAI
import os
import json
import re

nltk.download('punkt')
nltk.download('stopwords')

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'postgresql://postgres:Gied@localhost/DiversityDB'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = 'secret_key'  

db = SQLAlchemy(app)

# Set OpenAI API key using an environment variable
os.environ["OPENAI_API_KEY"] = os.getenv("OPENAI_API_KEY")

# Initialise OpenAI client
client = OpenAI(api_key=os.environ.get("OPENAI_API_KEY"))


class User(db.Model):
    __tablename__ = 'users'
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(120), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


# Load Lemmas from Excel File
def load_lemmas():
    workbook = openpyxl.load_workbook('Lemmas.xlsx', read_only=True)

    educational_lemmas = set(row[0].value.lower() for row in workbook['Educational'].iter_rows(min_row=2, max_col=1))
    social_lemmas = set(row[0].value.lower() for row in workbook['Social'].iter_rows(min_row=2, max_col=1))
    technical_lemmas = set(row[0].value.lower() for row in workbook['Technical'].iter_rows(min_row=2, max_col=1))

    return educational_lemmas, social_lemmas, technical_lemmas

educational_lemmas, social_lemmas, technical_lemmas = load_lemmas()


def preprocess_and_tokenize(text):
    words = word_tokenize(text.lower())
    return words


def clickableUrls(text):
    # Regular expression to detect URLs
    url_pattern = r'https?://\S+'
    # Replace URLs with clickable anchor tags
    return re.sub(url_pattern, lambda x: f"<a href='{x.group(0)}'>{x.group(0)}</a>", text)


def analyze_text(text):
    text = clickableUrls(text)
    words = preprocess_and_tokenize(text)

    educational_lemmas_lower = {lemma.lower() for lemma in educational_lemmas}
    social_lemmas_lower = {lemma.lower() for lemma in social_lemmas}
    technical_lemmas_lower = {lemma.lower() for lemma in technical_lemmas}

    educational_count = sum(1 for word in words if word.lower() in educational_lemmas_lower)
    social_count = sum(1 for word in words if word.lower() in social_lemmas_lower)
    technological_count = sum(1 for word in words if word.lower() in technical_lemmas_lower)

    total_lemma_words = educational_count + social_count + technological_count

    # Count total words in description
    total_words = len(text.split())

    # Calculate scores based on the proportion of each category in the description
    educational_score = round((educational_count / total_lemma_words) * 100, 2)
    social_score = round((social_count / total_lemma_words) * 100, 2)
    technological_score = round((technological_count / total_lemma_words) * 100, 2)

    modified_text = f"Print Recommendations."

    return educational_score, social_score, technological_score, educational_count, social_count, technological_count, total_words, modified_text


def filter_descriptions(courses):
    filtered_courses = [] # Filter out descriptions that have 50 or less words
    for course in courses:
        description = course['description']
        if isinstance(description, list):
            # Join the list of strings into a single string
            description = ' '.join(description)
        if isinstance(description, str):
            word_count = len(description.split())
            if word_count >= 50:
                filtered_courses.append(course)
    return filtered_courses


@app.route('/modify_description', methods=['POST'])
def modify_description():
    text_to_modify = request.form['text']
    modification_type = request.form['modification_type']

    # Analyze the text to get scores and counts
    educational_score, social_score, technological_score, educational_count, social_count, technological_count, _, total_words = analyze_text(text_to_modify)

    max_tokens_limit = 2500

    # Calculate the target count for the chosen category in the modified description
    if modification_type == 'educational':
        target_count = educational_count + 5
    elif modification_type == 'social':
        target_count = social_count + 5
    elif modification_type == 'technical':
        target_count = technological_count + 5
    else:
        # Handle invalid modification type
        return {'error': 'Invalid modification type'}

    # Retrieve the corresponding lemma set based on the modification type
    lemmas_set = {
        'educational': educational_lemmas,
        'social': social_lemmas,
        'technical': technical_lemmas,
    }.get(modification_type, set())

    # Extract a subset of lemmas to be used in the modification
    lemmas_to_use = list(lemmas_set)

    # Determine the length of the description modification
    max_length = len(text_to_modify.split())

    # Construct the assistant message content with lemmas
    assistant_message = f"Use {target_count} or more of the following lemmas in the description. Lemmas: ({', '.join(lemmas_to_use)}.)" \
                        f"Make the description more {modification_type} but keep original contents of the description. Modify to a length of approximately {max_length} words. " 
                        

    # API request to OpenAI for description modification
    try:
        response = client.chat.completions.create(
            model="gpt-3.5-turbo",
            messages=[
                {"role": "system", "content": "You are a helpful assistant."},
                {"role": "user", "content": text_to_modify},
                {"role": "assistant", "content": assistant_message},
            ],
            max_tokens=max_tokens_limit,
        )

        # Access the assistant's reply from 'choices' key
        modified_text = response.choices[0].message.content

        # Reanalyze the modified text to get updated scores and counts
        educational_score, social_score, technological_score, educational_count, social_count, technological_count, _, total_words = analyze_text(modified_text)

        # Highlight technical, social, and educational words in the modified text
        modified_text_with_highlights = ""
        for word in modified_text.split():
            clean_word = word.strip('.:,')
            if clean_word.lower() in educational_lemmas:
                modified_text_with_highlights += f"<span class='educational'>{word}</span> "
            elif clean_word.lower() in social_lemmas:
                modified_text_with_highlights += f"<span class='social'>{word}</span> "
            elif clean_word.lower() in technical_lemmas:
                modified_text_with_highlights += f"<span class='technological'>{word}</span> "
            else:
                modified_text_with_highlights += f"{word} "

    except Exception as e:
        # Handle errors
        modified_text_with_highlights = f"Error: {str(e)}"

    return {
        'modified_text': modified_text_with_highlights,
        'edu_score': educational_score,
        'social_score': social_score,
        'tech_score': technological_score,
        'edu_count': educational_count,
        'social_count': social_count,
        'tech_count': technological_count,
        'word_count': total_words
    }


@app.route('/login', methods=['POST', 'GET'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        print(f"Received username: {username}, password: {password}")

        user = User.query.filter_by(username=username).first()

        if user:
            print(f"Retrieved user: {user.username}, {user.password_hash}")

            # Verify the password using check_password_hash
            if check_password_hash(user.password_hash, password):
                session['user_id'] = user.id
                return redirect(url_for('index'))
        
        print("Invalid username or password.")
        error = 'Invalid username or password.'
        return render_template('login.html', error=error)

    return render_template('login.html')


@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form['username']
        email = request.form['email']
        password = request.form['password']

        hashed_password = generate_password_hash(password, method='pbkdf2:sha256')
        new_user = User(username=username, email=email, password_hash=hashed_password)

        db.session.add(new_user)
        db.session.commit()

        return redirect(url_for('login'))

    return render_template('register.html')


@app.route('/forgot_password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        username = request.form['username']
        old_password = request.form['old_password']
        new_password = request.form['new_password']
        confirm_password = request.form['confirm_password']

        # Check if the username exists
        user = User.query.filter_by(username=username).first()
        if not user:
            error = 'Username not found.'
            return render_template('forgot_password.html', error=error)

        # Verify the old password
        if not check_password_hash(user.password_hash, old_password):
            error = 'Incorrect old password.'
            return render_template('forgot_password.html', error=error)

        # Check if the new password and confirm password match
        if new_password != confirm_password:
            error = 'New password and confirm password do not match.'
            return render_template('forgot_password.html', error=error)

        # Update the password
        user.password_hash = generate_password_hash(new_password, method='pbkdf2:sha256')
        db.session.commit()

        return redirect(url_for('login'))

    return render_template('forgot_password.html')


@app.route('/logout')
def logout():
    session.pop('user_id', None)
    return redirect(url_for('login'))


@app.route('/')
def index():
    if 'user_id' in session:
        return render_template('index.html')
    else:
        return redirect(url_for('login'))


@app.route('/courses_json')
def courses_json():
    with open('ingenic_courses/courses.json', encoding='utf-8') as f:
        courses = json.load(f)
    
    filtered_courses = filter_descriptions(courses)
    
    return jsonify(filtered_courses)


@app.route('/analyze', methods=['POST'])
def analyze():
    text_to_analyze = request.form['text']

    educational_score, social_score, technological_score, educational_count, social_count, technological_count, total_words, modified_text = analyze_text(text_to_analyze)

    # Get the lists of educational, social, and technological words
    educational_words = list(educational_lemmas)
    social_words = list(social_lemmas)
    technological_words = list(technical_lemmas)

    return render_template('result.html', original_text=text_to_analyze, modified_text=modified_text,
                           edu_score=educational_score, social_score=social_score, tech_score=technological_score,
                           edu_count=educational_count, social_count=social_count, tech_count=technological_count,
                           total_words=total_words, educational_words=educational_words,
                           social_words=social_words, technological_words=technological_words)


if __name__ == '__main__':
    db.create_all()
    app.run(debug=True, use_reloader=True)